import asyncio
import logging
import os
from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

from dotenv import set_key
from fastapi import APIRouter, Body, Depends, Header, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import desc, func, nullslast

from api.auth import create_token, verify_token, clear_all_admin_tokens
from core.config import BASE_DIR
from api.auth import _hash_password as _hash
from core.database import AgentConfigORM, ButtonClickORM, CompanionAgentConfigORM, CompanionORM, CompanionStateORM, ConfigGroupORM, FeedbackMessageORM, FeedbackThreadORM, MomentORM, PageViewORM, ShortTermMessageORM, SystemNotificationORM, UserCompanionStateORM, UserORM, get_db, serialize_datetime
from core.state import get_companion_manager
from services.agent import test_llm_connection
from services.analytics_stats import compute_dau_series, compute_retention_cohorts
from services.knowledge_base import knowledge_base
from services.memory import get_embedding_status
from services.culture_data import (
    BATCH_GENERATION_ALL_LANGS_ORDER,
    build_batch_profiles,
    get_batch_persona_output_instruction,
    get_cultural_context,
    infer_language_from_city,
    normalize_batch_generation_lang,
)
from services.image_generation import generate_image_with_cache
from services.async_tasks import start_avatar_generation, start_moment_image_generation

router = APIRouter()
logger = logging.getLogger(__name__)


# ===== 后端错误信息多语言 =====
_ADMIN_ERROR_MESSAGES = {
    "zh": {
        "missing_auth": "缺少认证信息",
        "invalid_token": "Token 无效或已过期",
        "wrong_password": "密码错误",
        "companion_not_found": "智能体不存在",
        "entry_not_found": "条目不存在",
        "moment_not_found": "朋友圈不存在",
        "moment_no_caption": "朋友圈不存在或无文案",
        "config_group_not_found": "配置组不存在",
        "key_name_required": "key 和 name 不能为空",
        "invalid_config_type": "config_type 必须是 model_service、agent 或 image_generation",
        "key_exists": "key 已存在",
        "feedback_not_found": "反馈会话不存在",
        "user_not_found": "用户不存在",
        "notification_not_found": "通知不存在",
        "batch_generate_failed": "批量生成失败",
        "companion_id_caption_required": "companion_id 和 caption 为必填项",
        "empty_content": "回复内容不能为空",
        "username_too_short": "用户名至少3个字符",
        "password_too_short": "密码至少6个字符",
        "username_exists": "用户名已存在",
        "prompt_required": "prompt 为必填项",
        "title_content_required": "标题和内容不能为空",
        "unsupported_language": "不支持的语言",
        "server_error": "服务器内部错误",
        "reset_confirm_required": "请确认清空全部数据，confirm 必须为 yes_clear_all_data",
        "invalid_affection": "亲密度必须在 0～100 之间",
        "affection_required": "请填写亲密度",
        "session_not_found": "未找到该用户与智能体的会话记录",
    },
    "en": {
        "missing_auth": "Missing authentication",
        "invalid_token": "Invalid or expired token",
        "wrong_password": "Incorrect password",
        "companion_not_found": "Companion not found",
        "entry_not_found": "Entry not found",
        "moment_not_found": "Moment not found",
        "moment_no_caption": "Moment not found or has no caption",
        "config_group_not_found": "Config group not found",
        "key_name_required": "key and name are required",
        "invalid_config_type": "config_type must be model_service, agent or image_generation",
        "key_exists": "key already exists",
        "feedback_not_found": "Feedback thread not found",
        "user_not_found": "User not found",
        "notification_not_found": "Notification not found",
        "batch_generate_failed": "Batch generation failed",
        "companion_id_caption_required": "companion_id and caption are required",
        "empty_content": "Reply content cannot be empty",
        "username_too_short": "Username must be at least 3 characters",
        "password_too_short": "Password must be at least 6 characters",
        "username_exists": "Username already exists",
        "prompt_required": "prompt is required",
        "title_content_required": "Title and content are required",
        "unsupported_language": "Unsupported language",
        "server_error": "Internal server error",
        "reset_confirm_required": "Please confirm to clear all data, confirm must be yes_clear_all_data",
        "invalid_affection": "Affection must be between 0 and 100",
        "affection_required": "Affection is required",
        "session_not_found": "No chat session for this user and companion",
    },
}


def _get_error_msg(key: str, lang: str = "zh") -> str:
    """根据语言获取错误信息"""
    lang = lang.split("-")[0]  # zh-CN -> zh
    return _ADMIN_ERROR_MESSAGES.get(lang, _ADMIN_ERROR_MESSAGES["zh"]).get(key, key)


async def get_admin_lang(accept_language: Optional[str] = Header(None, alias="Accept-Language")) -> str:
    """从请求头解析管理员界面语言，默认中文"""
    if accept_language:
        lang = accept_language.split(",")[0].strip().lower()
        lang = lang.split("-")[0]
        if lang in _ADMIN_ERROR_MESSAGES:
            return lang
    return "zh"


async def admin_auth_required(authorization: str = Header(...)):
    if not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing authentication")
    token = authorization[7:]
    if not verify_token(token):
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return token


@router.post("/api/admin/login")
async def admin_login(data: dict, lang: str = Depends(get_admin_lang)):
    password = data.get("password", "")
    try:
        token = create_token(password)
    except ValueError:
        raise HTTPException(status_code=500, detail=_get_error_msg("server_error", lang))
    if not token:
        raise HTTPException(status_code=401, detail=_get_error_msg("wrong_password", lang))
    return {"token": token}


@router.get("/api/admin/stats")
async def admin_stats(_token: str = Depends(admin_auth_required)):
    companions = get_companion_manager().list_all_for_admin()
    total_turns = sum(c["state"].get("turns", 0) for c in companions)
    avg_affection = sum(c["state"].get("affection", 0) for c in companions) / max(len(companions), 1)
    return {
        "companion_count": len(companions),
        "total_turns": total_turns,
        "avg_affection": round(avg_affection, 1),
        "knowledge_stats": knowledge_base.get_stats(),
    }


@router.get("/api/admin/companions")
async def admin_list_companions(_token: str = Depends(admin_auth_required)):
    return get_companion_manager().list_all_for_any()


@router.delete("/api/admin/companions/{companion_id}")
async def admin_delete_companion(companion_id: str, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    ok = get_companion_manager().delete(companion_id)
    if not ok:
        raise HTTPException(status_code=404, detail=_get_error_msg("companion_not_found", lang))
    return {"ok": True}


@router.put("/api/admin/companions/{companion_id}")
async def admin_update_companion(companion_id: str, data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    companion = get_companion_manager().update(companion_id, data)
    if not companion:
        raise HTTPException(status_code=404, detail=_get_error_msg("companion_not_found", lang))
    return companion.to_dict()


@router.get("/api/admin/companions/{companion_id}")
async def admin_get_companion(companion_id: str, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """获取单个伴侣详情（供编辑表单使用），合并 profile 和 agent-config 中的 system_prompt"""
    cm = get_companion_manager()
    companion = cm.get(companion_id)
    if not companion:
        # 尝试从 DB 加载（内存中可能未加载）
        with get_db() as db:
            row = db.query(CompanionORM).filter(CompanionORM.id == companion_id).first()
            if not row:
                raise HTTPException(status_code=404, detail=_get_error_msg("companion_not_found", lang))
            # 触发加载到内存
            companion = cm.get_or_create(companion_id)
            if not companion:
                raise HTTPException(status_code=404, detail=_get_error_msg("companion_not_found", lang))

    data = companion.to_dict() if hasattr(companion, 'to_dict') else {}

    # 合并 agent-config 中的 system_prompt_*
    with get_db() as db:
        config_row = db.query(CompanionAgentConfigORM).filter(
            CompanionAgentConfigORM.companion_id == companion_id
        ).first()
        if config_row and config_row.config_json:
            cfg = config_row.config_json
            for k, v in cfg.items():
                if k.startswith("system_prompt_") or k in ("temperature", "max_tokens"):
                    data[k] = v
            # 也把 profile 显式添加（前端期望）
            if "profile" not in data:
                data["profile"] = {
                    k: v for k, v in data.items() if not k.startswith("system_prompt_")
                }

    return data


@router.get("/api/admin/knowledge")
async def admin_list_knowledge(category: str = None, language: str = None, _token: str = Depends(admin_auth_required)):
    return knowledge_base.list_entries(category=category, language=language)


@router.post("/api/admin/knowledge")
async def admin_add_knowledge(data: dict, _token: str = Depends(admin_auth_required)):
    entry = knowledge_base.add_entry(data)
    return entry.model_dump()


@router.delete("/api/admin/knowledge/{entry_id}")
async def admin_delete_knowledge(entry_id: str, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    ok = knowledge_base.delete_entry(entry_id)
    if not ok:
        raise HTTPException(status_code=404, detail=_get_error_msg("entry_not_found", lang))
    return {"ok": True}


@router.get("/api/admin/moments")
async def admin_list_moments(limit: int = 50, offset: int = 0, lang: Optional[str] = None, _token: str = Depends(admin_auth_required)):
    from services.moments import get_moments_feed
    # 传递 lang 参数，确保管理端列表本地化与用户端一致（修复之前未传 lang 的问题）
    effective_lang = lang or "zh"
    moments = get_moments_feed(limit=limit, offset=offset, device_id="", lang=effective_lang)
    with get_db() as db:
        total = db.query(func.count(MomentORM.id)).scalar() or 0
    return {"moments": moments, "total": total}


@router.delete("/api/admin/moments")
async def admin_clear_all_moments(_token: str = Depends(admin_auth_required)):
    """清空所有朋友圈（包括点赞和评论）"""
    from services.moments import clear_all_moments
    count = clear_all_moments()
    return {"ok": True, "deleted": count}


@router.post("/api/admin/reset-all")
async def admin_reset_all(data: Optional[dict] = None, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """一键清除全部数据（所有智能体、朋友圈、知识库、记忆文件等）。
    必须提供 confirm=yes_clear_all_data 才能执行，避免误操作。
    可选参数 regenerate=true 会在清空后尝试重新生成朋友圈（需先通过 batch-generate 创建智能体）。
    """
    if not data or data.get("confirm") != "yes_clear_all_data":
        raise HTTPException(
            status_code=400, 
            detail=_get_error_msg("reset_confirm_required", lang)
        )

    cm = get_companion_manager()
    deleted = cm.clear_all()

    # 清空其他非核心数据（用户、反馈、统计等，可根据需要保留配置）
    with get_db() as db:
        db.query(UserORM).delete(synchronize_session=False)
        db.query(FeedbackThreadORM).delete(synchronize_session=False)
        db.query(FeedbackMessageORM).delete(synchronize_session=False)
        db.query(PageViewORM).delete(synchronize_session=False)
        db.query(ButtonClickORM).delete(synchronize_session=False)
        db.query(SystemNotificationORM).delete(synchronize_session=False)
        db.commit()

    regenerate = data.get("regenerate", False)
    moments_created = 0
    if regenerate:
        try:
            from services.moments import regenerate_moments_for_all
            import asyncio
            loop = asyncio.get_event_loop()
            # 注意：清空后需先创建智能体，否则 moments_created=0
            moments_created = await loop.run_in_executor(
                None, regenerate_moments_for_all, cm, data.get("moments_per_companion", 3)
            )
        except Exception as e:
            logger.warning("[Reset] 重新生成朋友圈失败: %s", e)

    return {
        "ok": True,
        "deleted_companions": deleted,
        "deleted_moments_related": "all",
        "moments_regenerated": moments_created,
        "regenerate": regenerate,
        "message": "全部数据已成功清除" + ("并尝试重新生成朋友圈" if regenerate else "。请使用批量生成接口重新创建智能体"),
        "next_step": "推荐运行 python server/scripts/batch_create_all_langs.py 重新生成智能体，然后调用 /moments/regenerate"
    }


@router.post("/api/admin/moments/regenerate")
async def admin_regenerate_moments(data: Optional[dict] = None, _token: str = Depends(admin_auth_required)):
    """清空所有历史朋友圈并重新生成，可选参数 moments_per_companion 控制每个伴侣生成数量"""
    from services.moments import regenerate_moments_for_all
    import asyncio
    cm = get_companion_manager()
    moments_per = 3
    if data and isinstance(data, dict):
        moments_per = data.get("moments_per_companion", 3)
    loop = asyncio.get_event_loop()
    created = await loop.run_in_executor(None, regenerate_moments_for_all, cm, moments_per)
    return {"ok": True, "created": created}


@router.post("/api/admin/moments/batch-generate")
async def admin_batch_generate_moments(data: dict, _token: str = Depends(admin_auth_required)):
    """一键批量生成朋友圈，根据人设+文案自动生成配图。
    参数:
      - clear_existing: bool 是否清空现有朋友圈（默认 false，追加生成）
      - moments_per_companion: int 每个伴侣生成数量（默认 3）
    """
    from services.moments import regenerate_moments_for_all, create_moment_for_companion
    import asyncio
    cm = get_companion_manager()
    clear_existing = data.get("clear_existing", False)
    moments_per = data.get("moments_per_companion", 3)
    loop = asyncio.get_event_loop()

    if clear_existing:
        created = await loop.run_in_executor(None, regenerate_moments_for_all, cm, moments_per)
        return {"ok": True, "created": created, "mode": "clear_and_regenerate"}

    # 追加生成模式：不清空，为每个伴侣追加生成
    created = 0
    companions = cm.list_all()
    for c in companions:
        companion_id = c.get("profile", {}).get("id")
        if not companion_id:
            continue
        companion = cm.get(companion_id)
        if not companion:
            continue
        for _ in range(moments_per):
            result = await loop.run_in_executor(None, create_moment_for_companion, companion, cm)
            if result:
                created += 1
    return {"ok": True, "created": created, "mode": "append"}


@router.delete("/api/admin/moments/{moment_id}")
async def admin_delete_moment(moment_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    from services.moments import delete_moment
    ok = delete_moment(moment_id)
    if not ok:
        raise HTTPException(status_code=404, detail=_get_error_msg("moment_not_found", lang))
    return {"ok": True}


@router.post("/api/admin/moments/{moment_id}/regenerate-image")
async def admin_regenerate_moment_image(moment_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """重新生成指定朋友圈的配图"""
    from services.moments import regenerate_moment_image
    new_url = regenerate_moment_image(moment_id)
    if not new_url:
        raise HTTPException(status_code=404, detail=_get_error_msg("moment_no_caption", lang))
    return {"ok": True, "image_url": new_url}


@router.get("/api/admin/users")
async def admin_list_users(_token: str = Depends(admin_auth_required)):
    from core.database import UserORM
    with get_db() as db:
        users = db.query(UserORM).order_by(UserORM.created_at.desc()).all()
        return [
            {
                "id": u.id,
                "username": u.username,
                "nickname": u.nickname,
                "gender": u.gender,
                "sexual_orientation": u.sexual_orientation,
                "age": getattr(u, "age", None),
                "region": (u.region or "") if getattr(u, "region", None) is not None else "",
                "occupation": (u.occupation or "") if getattr(u, "occupation", None) is not None else "",
                "created_at": u.created_at.isoformat() if u.created_at else None,
            }
            for u in users
        ]


@router.post("/api/admin/knowledge/search")
async def admin_search_knowledge(data: dict, _token: str = Depends(admin_auth_required)):
    query = data.get("query", "")
    top_k = data.get("top_k", 10)
    if not query:
        return {"results": []}
    results = knowledge_base.search_entries(query, top_k=top_k)
    return {"results": results}


@router.get("/api/admin/embedding-status")
async def admin_embedding_status(_token: str = Depends(admin_auth_required)):
    return get_embedding_status()


# ===== 配置组管理 =====

@router.get("/api/admin/config-groups")
async def admin_list_config_groups(_token: str = Depends(admin_auth_required)):
    with get_db() as db:
        rows = db.query(ConfigGroupORM).order_by(ConfigGroupORM.sort_order.asc(), ConfigGroupORM.id.asc()).all()
        return [
            {
                "id": r.id,
                "key": r.key,
                "name": r.name,
                "description": r.description,
                "config_type": r.config_type,
                "config_json": r.config_json or {},
                "enabled": bool(r.enabled),
                "sort_order": r.sort_order,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            }
            for r in rows
        ]


@router.get("/api/admin/config-groups/{group_id}")
async def admin_get_config_group(group_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        row = db.query(ConfigGroupORM).filter(ConfigGroupORM.id == group_id).first()
        if not row:
            raise HTTPException(status_code=404, detail=_get_error_msg("config_group_not_found", lang))
        return {
            "id": row.id,
            "key": row.key,
            "name": row.name,
            "description": row.description,
            "config_type": row.config_type,
            "config_json": row.config_json or {},
            "enabled": bool(row.enabled),
            "sort_order": row.sort_order,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        }


@router.post("/api/admin/config-groups")
async def admin_create_config_group(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    key = data.get("key", "").strip()
    name = data.get("name", "").strip()
    config_type = data.get("config_type", "agent").strip()
    if not key or not name:
        raise HTTPException(status_code=400, detail=_get_error_msg("key_name_required", lang))
    if config_type not in ("model_service", "agent", "image_generation"):
        raise HTTPException(status_code=400, detail=_get_error_msg("invalid_config_type", lang))
    with get_db() as db:
        exists = db.query(ConfigGroupORM).filter(ConfigGroupORM.key == key).first()
        if exists:
            raise HTTPException(status_code=400, detail=_get_error_msg("key_exists", lang))
        default_config = {}
        if config_type == "model_service":
            default_config = {
                "model_provider": "anthropic",
                "anthropic_ready": False,
                "deepseek_ready": False,
                "openai_ready": False,
                "admin_password_set": False,
            }
        elif config_type == "agent":
            default_config = {
                "model_provider": "",
                "temperature": 0.93,
                "max_tokens": 2048,
                "system_prompt_zh": "",
                "system_prompt_en": "",
                "system_prompt_ja": "",
                "system_prompt_ko": "",
                "system_prompt_pt": "",
                "system_prompt_es": "",
                "system_prompt_id": "",
            }
        elif config_type == "image_generation":
            default_config = {
                "provider": "volcano",
                "base_url": "https://ark.cn-beijing.volces.com/api/v3/images/generations",
                "api_key": "",
                "model": "doubao-seedream-5-0-260128",
                "size": "2K",
                "default_width": 1024,
                "default_height": 1024,
            }
        group = ConfigGroupORM(
            key=key,
            name=name,
            description=data.get("description", ""),
            config_type=config_type,
            config_json={**default_config, **(data.get("config_json") or {})},
            enabled=1 if data.get("enabled", True) else 0,
            sort_order=data.get("sort_order", 0),
        )
        db.add(group)
        db.flush()
        return {
            "id": group.id,
            "key": group.key,
            "name": group.name,
            "config_type": group.config_type,
            "enabled": bool(group.enabled),
        }


@router.put("/api/admin/config-groups/{group_id}")
async def admin_update_config_group(group_id: int, data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        row = db.query(ConfigGroupORM).filter(ConfigGroupORM.id == group_id).first()
        if not row:
            raise HTTPException(status_code=404, detail=_get_error_msg("config_group_not_found", lang))
        if "enabled" in data:
            row.enabled = 1 if data["enabled"] else 0
        if "name" in data:
            row.name = data["name"]
        if "description" in data:
            row.description = data["description"]
        if "sort_order" in data:
            row.sort_order = data["sort_order"]
        if "config_json" in data:
            row.config_json = data["config_json"]
    return {"ok": True}


@router.delete("/api/admin/config-groups/{group_id}")
async def admin_delete_config_group(group_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        row = db.query(ConfigGroupORM).filter(ConfigGroupORM.id == group_id).first()
        if not row:
            raise HTTPException(status_code=404, detail=_get_error_msg("config_group_not_found", lang))
        db.delete(row)
    return {"ok": True}


@router.get("/api/admin/config")
async def admin_config(_token: str = Depends(admin_auth_required)):
    return {
        "anthropic_ready": bool(os.getenv("ANTHROPIC_API_KEY", "")),
        "openai_ready": bool(os.getenv("OPENAI_API_KEY", "")),
        "deepseek_ready": bool(os.getenv("DEEPSEEK_API_KEY", "")),
        "model_provider": os.getenv("MODEL_PROVIDER", "anthropic"),
        "admin_password_set": bool(os.getenv("ADMIN_PASSWORD", "")),
    }


@router.post("/api/admin/config")
async def admin_update_config(data: dict, _token: str = Depends(admin_auth_required)):
    dotenv_path = str(BASE_DIR / "server" / ".env")
    updated = []

    field_map = {
        "anthropic_key": "ANTHROPIC_API_KEY",
        "openai_key": "OPENAI_API_KEY",
        "deepseek_key": "DEEPSEEK_API_KEY",
        "admin_password": "ADMIN_PASSWORD",
    }

    for key, env_name in field_map.items():
        value = data.get(key)
        if value is not None:
            set_key(dotenv_path, env_name, value)
            os.environ[env_name] = value
            updated.append(env_name)

    provider = data.get("model_provider")
    if provider is not None:
        set_key(dotenv_path, "MODEL_PROVIDER", provider)
        os.environ["MODEL_PROVIDER"] = provider
        updated.append("MODEL_PROVIDER")

    if "ADMIN_PASSWORD" in updated:
        clear_all_admin_tokens()

    return {"ok": True, "updated": updated}


@router.post("/api/admin/config/test")
async def admin_test_config(data: dict, _token: str = Depends(admin_auth_required)):
    provider = data.get("provider") or os.getenv("MODEL_PROVIDER", "anthropic")
    result = test_llm_connection(provider=provider)
    return result


@router.post("/api/admin/config-groups/{group_id}/test")
async def admin_test_config_group(
    group_id: int,
    data: Optional[dict] = Body(None),
    _token: str = Depends(admin_auth_required),
    lang: str = Depends(get_admin_lang),
):
    """根据配置组类型执行连通性测试"""
    with get_db() as db:
        row = db.query(ConfigGroupORM).filter(ConfigGroupORM.id == group_id).first()
        if not row:
            raise HTTPException(status_code=404, detail=_get_error_msg("config_group_not_found", lang))

        cfg = row.config_json or {}
        config_type = row.config_type

        if config_type == "model_service":
            data = data or {}
            provider = data.get("provider") or cfg.get("model_provider") or os.getenv("MODEL_PROVIDER", "anthropic")
            cfg_eff = dict(cfg)
            for k in ("anthropic_key", "deepseek_key", "openai_key"):
                v = data.get(k)
                if isinstance(v, str) and v.strip():
                    cfg_eff[k] = v.strip()
            overrides = {}
            for k in ("anthropic_key", "deepseek_key", "openai_key"):
                v = cfg_eff.get(k)
                if isinstance(v, str) and v.strip():
                    overrides[k] = v.strip()
            result = test_llm_connection(provider=provider, api_key_overrides=overrides if overrides else None)
            return result

        if config_type == "image_generation":
            provider = cfg.get("provider", "")
            base_url = cfg.get("base_url", "")
            api_key = cfg.get("api_key", "")
            access_key_id = cfg.get("access_key_id", "")
            secret_access_key = cfg.get("secret_access_key", "")
            session_token = cfg.get("session_token", "")
            if not provider:
                return {"ok": False, "error": "Provider 未配置"}
            if provider.lower() != "volcano":
                return {"ok": False, "error": "仅支持火山引擎 (volcano) provider"}
            if not api_key and not (access_key_id and secret_access_key):
                return {"ok": False, "error": "API Key 或 AK/SK 未配置"}
            if provider.lower() == "volcano":
                test_url = (base_url or "https://ark.cn-beijing.volces.com/api/v3").rstrip("/") + "/models"
                try:
                    import requests
                    if api_key:
                        resp = requests.get(
                            test_url,
                            headers={"Authorization": f"Bearer {api_key}"},
                            timeout=10,
                        )
                    else:
                        # 使用 AK/SK 签名鉴权
                        from services.image_generation import _sign_volcano_request
                        body_bytes = b""
                        headers = _sign_volcano_request(
                            method="GET",
                            url=test_url,
                            body=body_bytes,
                            access_key=access_key_id,
                            secret_key=secret_access_key,
                            session_token=session_token,
                        )
                        resp = requests.get(test_url, headers=headers, timeout=10)
                    resp.raise_for_status()
                    return {"ok": True, "response": "火山引擎 API 连接正常"}
                except Exception as e:
                    return {"ok": False, "error": str(e)}
            return {"ok": True, "response": "配置已设置"}

        if config_type == "agent":
            override_provider = cfg.get("model_provider", "")
            if override_provider:
                result = test_llm_connection(provider=override_provider)
                return result
            temperature = cfg.get("temperature")
            max_tokens = cfg.get("max_tokens")
            if temperature is None or max_tokens is None:
                return {"ok": False, "error": "Temperature 或 Max Tokens 未配置"}
            return {"ok": True, "response": "配置完整"}

        return {"ok": False, "error": f"未知配置类型: {config_type}"}


@router.get("/api/admin/agent-config")
async def admin_get_agent_config(_token: str = Depends(admin_auth_required)):
    with get_db() as db:
        row = db.query(AgentConfigORM).first()
        if row:
            return row.config_json or {}
    return {}


@router.put("/api/admin/agent-config")
async def admin_put_agent_config(data: dict, _token: str = Depends(admin_auth_required)):
    with get_db() as db:
        row = db.query(AgentConfigORM).first()
        if row:
            cfg = dict(row.config_json or {})
            cfg.update(data)
            row.config_json = cfg
        else:
            db.add(AgentConfigORM(config_json=data))
    return {"ok": True}


@router.get("/api/admin/companions/{companion_id}/agent-config")
async def admin_get_companion_agent_config(companion_id: str, _token: str = Depends(admin_auth_required)):
    with get_db() as db:
        row = db.query(CompanionAgentConfigORM).filter(
            CompanionAgentConfigORM.companion_id == companion_id
        ).first()
        if row and row.config_json:
            return row.config_json
    return {}


@router.put("/api/admin/companions/{companion_id}/agent-config")
async def admin_put_companion_agent_config(companion_id: str, data: dict, _token: str = Depends(admin_auth_required)):
    with get_db() as db:
        row = db.query(CompanionAgentConfigORM).filter(
            CompanionAgentConfigORM.companion_id == companion_id
        ).first()
        if row:
            cfg = dict(row.config_json or {})
            cfg.update(data)
            row.config_json = cfg
        else:
            db.add(CompanionAgentConfigORM(companion_id=companion_id, config_json=data))
    return {"ok": True}


@router.get("/api/admin/feedback")
async def admin_list_feedback(_token: str = Depends(admin_auth_required)):
    with get_db() as db:
        threads = db.query(FeedbackThreadORM).order_by(
            FeedbackThreadORM.updated_at.desc()
        ).all()

        # 批量查询最后一条消息，避免 N+1
        thread_ids = [t.id for t in threads]
        last_msg_map = {}
        if thread_ids:
            messages = (
                db.query(FeedbackMessageORM)
                .filter(FeedbackMessageORM.thread_id.in_(thread_ids))
                .order_by(FeedbackMessageORM.thread_id, FeedbackMessageORM.created_at.desc())
                .all()
            )
            for m in messages:
                if m.thread_id not in last_msg_map:
                    last_msg_map[m.thread_id] = m

        return [
            {
                "id": t.id,
                "user_id": t.user_id,
                "user_name": t.user_name,
                "status": t.status,
                "created_at": t.created_at.isoformat() if t.created_at else None,
                "updated_at": t.updated_at.isoformat() if t.updated_at else None,
                "last_message": last_msg_map.get(t.id).content if last_msg_map.get(t.id) else "",
                "last_message_sender": last_msg_map.get(t.id).sender if last_msg_map.get(t.id) else "",
            }
            for t in threads
        ]


@router.get("/api/admin/feedback/{thread_id}/messages")
async def admin_get_feedback_messages(thread_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        thread = db.query(FeedbackThreadORM).filter(
            FeedbackThreadORM.id == thread_id
        ).first()
        if not thread:
            raise HTTPException(status_code=404, detail=_get_error_msg("feedback_not_found", lang))

        messages = db.query(FeedbackMessageORM).filter(
            FeedbackMessageORM.thread_id == thread_id
        ).order_by(FeedbackMessageORM.created_at.asc()).all()

        return {
            "thread": {
                "id": thread.id,
                "user_id": thread.user_id,
                "user_name": thread.user_name,
                "status": thread.status,
            },
            "messages": [
                {
                    "id": m.id,
                    "sender": m.sender,
                    "content": m.content,
                    "created_at": m.created_at.isoformat() if m.created_at else None,
                }
                for m in messages
            ],
        }


@router.get("/api/admin/chat-sessions/messages")
async def admin_get_chat_session_messages(
    user_id: int,
    companion_id: str,
    limit: int = 300,
    offset: int = 0,
    _token: str = Depends(admin_auth_required),
    lang: str = Depends(get_admin_lang),
):
    """某用户与某智能体对应的短期记忆消息（按智能体存储；与 App 端一致）。"""
    cid = (companion_id or "").strip()[:8]
    if not cid:
        raise HTTPException(status_code=400, detail=_get_error_msg("session_not_found", lang))

    limit = min(500, max(1, limit))
    offset = max(0, offset)

    with get_db() as db:
        ucs = (
            db.query(UserCompanionStateORM)
            .filter(
                UserCompanionStateORM.user_id == user_id,
                UserCompanionStateORM.companion_id == cid,
            )
            .first()
        )
        if not ucs:
            raise HTTPException(status_code=404, detail=_get_error_msg("session_not_found", lang))

        u = db.query(UserORM).filter(UserORM.id == user_id).first()
        comp = db.query(CompanionORM).filter(CompanionORM.id == cid).first()

        total = (
            db.query(ShortTermMessageORM)
            .filter(ShortTermMessageORM.companion_id == cid)
            .count()
        )

        msgs = (
            db.query(ShortTermMessageORM)
            .filter(ShortTermMessageORM.companion_id == cid)
            .order_by(ShortTermMessageORM.id.asc())
            .offset(offset)
            .limit(limit)
            .all()
        )

        # 必须在 session 内把字段拷成普通值再返回：get_db 退出时会 commit 并使 ORM 过期，
        # 在 with 外访问 msgs/u 会触发 DetachedInstanceError → 500 服务器内部错误。
        return {
            "user_id": user_id,
            "username": (u.username if u else "") or "",
            "nickname": (u.nickname if u else "") or "",
            "companion_id": cid,
            "companion_name": (comp.name if comp else "") or "",
            "total": int(total),
            "offset": offset,
            "messages": [
                {
                    "id": m.id,
                    "role": m.role,
                    "content": m.content or "",
                    "created_at": serialize_datetime(m.created_at),
                }
                for m in msgs
            ],
        }


@router.get("/api/admin/chat-sessions")
async def admin_list_chat_sessions(
    _token: str = Depends(admin_auth_required),
    page: int = 1,
    page_size: int = 20,
    user_id: Optional[int] = None,
    companion_id: Optional[str] = None,
):
    """IM 会话列表：用户与智能体维度（数据来自 user_companion_states）。"""
    page = max(1, page)
    page_size = min(100, max(1, page_size))
    offset = (page - 1) * page_size

    filters = []
    if user_id is not None:
        filters.append(UserCompanionStateORM.user_id == user_id)
    if companion_id:
        cid = companion_id.strip()[:8]
        if cid:
            filters.append(UserCompanionStateORM.companion_id == cid)

    with get_db() as db:
        total_q = db.query(UserCompanionStateORM)
        if filters:
            total_q = total_q.filter(*filters)
        total = total_q.count()

        list_q = (
            db.query(UserCompanionStateORM, UserORM, CompanionORM)
            .outerjoin(UserORM, UserORM.id == UserCompanionStateORM.user_id)
            .outerjoin(CompanionORM, CompanionORM.id == UserCompanionStateORM.companion_id)
        )
        if filters:
            list_q = list_q.filter(*filters)
        rows = (
            list_q.order_by(nullslast(desc(UserCompanionStateORM.updated_at)))
            .offset(offset)
            .limit(page_size)
            .all()
        )

        cids = list({ucs.companion_id for ucs, _, _ in rows})
        msg_stats: dict[str, tuple[int, Optional[datetime]]] = {}
        if cids:
            stats = (
                db.query(
                    ShortTermMessageORM.companion_id,
                    func.count(ShortTermMessageORM.id).label("cnt"),
                    func.max(ShortTermMessageORM.created_at).label("last_at"),
                )
                .filter(ShortTermMessageORM.companion_id.in_(cids))
                .group_by(ShortTermMessageORM.companion_id)
                .all()
            )
            for s in stats:
                msg_stats[s.companion_id] = (int(s.cnt or 0), s.last_at)

        items = []
        for ucs, u, c in rows:
            mc, last_at = msg_stats.get(ucs.companion_id, (0, None))
            items.append({
                "user_id": ucs.user_id,
                "username": (u.username if u else None) or "",
                "nickname": (u.nickname if u else None) or "",
                "companion_id": ucs.companion_id,
                "companion_name": (c.name if c else None) or "",
                "turns": ucs.turns if ucs.turns is not None else 0,
                "affection": float(ucs.affection) if ucs.affection is not None else 0.0,
                "session_updated_at": serialize_datetime(ucs.updated_at),
                "message_count": mc,
                "last_message_at": serialize_datetime(last_at),
            })

    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.delete("/api/admin/chat-sessions")
async def admin_delete_chat_session(
    user_id: int = Query(..., ge=1, description="用户 ID"),
    companion_id: str = Query(..., min_length=1, max_length=32, description="智能体 ID"),
    clear_messages: bool = Query(
        True,
        description="为 true 时同时清空该智能体的短期消息并重置亲密度（共用智能体时影响所有用户）",
    ),
    _token: str = Depends(admin_auth_required),
    lang: str = Depends(get_admin_lang),
):
    """删除一条用户–智能体会话记录；可选清空该智能体的短期 IM 文案。"""
    cid = (companion_id or "").strip()[:8]
    if not cid:
        raise HTTPException(status_code=400, detail=_get_error_msg("session_not_found", lang))

    with get_db() as db:
        ucs = (
            db.query(UserCompanionStateORM)
            .filter(
                UserCompanionStateORM.user_id == user_id,
                UserCompanionStateORM.companion_id == cid,
            )
            .first()
        )
        if not ucs:
            raise HTTPException(status_code=404, detail=_get_error_msg("session_not_found", lang))

        if clear_messages:
            db.query(ShortTermMessageORM).filter(
                ShortTermMessageORM.companion_id == cid
            ).delete(synchronize_session=False)
            gr = (
                db.query(CompanionStateORM)
                .filter(CompanionStateORM.companion_id == cid)
                .first()
            )
            if gr:
                gr.affection = 0.0
                gr.turns = 0
            db.query(UserCompanionStateORM).filter(
                UserCompanionStateORM.companion_id == cid
            ).update(
                {
                    UserCompanionStateORM.affection: 0,
                    UserCompanionStateORM.turns: 0,
                },
                synchronize_session=False,
            )

        db.query(UserCompanionStateORM).filter(
            UserCompanionStateORM.user_id == user_id,
            UserCompanionStateORM.companion_id == cid,
        ).delete(synchronize_session=False)

    companion = get_companion_manager().get(cid)
    if companion and clear_messages:
        companion.memory.short_term.clear()
        companion.state.affection = 0
        companion.state.turns = 0
        companion.save_state(user_id=None)

    return {"ok": True, "cleared_messages": clear_messages}


@router.post("/api/admin/feedback/{thread_id}/reply")
async def admin_reply_feedback(thread_id: int, data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    content = (data.get("content") or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail=_get_error_msg("empty_content", lang))

    with get_db() as db:
        thread = db.query(FeedbackThreadORM).filter(
            FeedbackThreadORM.id == thread_id
        ).first()
        if not thread:
            raise HTTPException(status_code=404, detail=_get_error_msg("feedback_not_found", lang))

        msg = FeedbackMessageORM(
            thread_id=thread_id,
            sender="admin",
            content=content,
        )
        db.add(msg)

        thread.status = "replied"

    return {"ok": True}


# ===== Companions: Create =====

@router.post("/api/admin/companions")
async def admin_create_companion(data: dict, _token: str = Depends(admin_auth_required)):
    try:
        companion = get_companion_manager().create(data)
        # 启动后台异步生成头像
        if companion.profile.avatar_url == "__GENERATING__":
            start_avatar_generation(companion.profile.id, companion.profile.model_dump())
        return companion.to_dict()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ===== Companions: Batch Generate =====

# ===== 批量生成人设的 fallback 文案（按语言） =====
_BATCH_FALLBACK_TEMPLATES = {
    "zh": {
        "background": "{name}来自{city}，是一位{age}岁的{gender}性，性格{personality}，有着丰富而独特的人生经历与内心世界。",
        "speech_style": "说话风格自然真实，充满{personality}的特质，语气亲切而富有个性。",
    },
    "en": {
        "background": "{name} is from {city}, a {age}-year-old {gender} with a {personality} personality, possessing a rich and unique life experience and inner world.",
        "speech_style": "Speaks in a natural and authentic style, full of {personality} traits, with a warm and distinctive tone.",
    },
    "ja": {
        "background": "{name}は{city}出身で、{age}歳の{gender}性です。性格は{personality}で、豊かで独特な人生経験と内面世界を持っています。",
        "speech_style": "自然でリアルな話し方で、{personality}の特質に満ち、親しみやすく個性的な口調です。",
    },
    "ko": {
        "background": "{name}은(는) {city} 출신의 {age}세 {gender}로, {personality}한 성격을 가지며 풍부하고 독특한 인생 경험과 깊이 있는 내면을 가지고 있습니다.",
        "speech_style": "자연스럽고 진실된 말투로, {personality}한 특성이 가득하며 친근하고 개성 넘치는 어조입니다.",
    },
    "pt": {
        "background": "{name} é de {city}, uma pessoa de {age} anos, do sexo {gender}, com personalidade {personality}, possuindo uma rica e única experiência de vida e mundo interior.",
        "speech_style": "Estilo de fala natural e autêntico, cheio de traços {personality}, com um tom caloroso e distinto.",
    },
    "es": {
        "background": "{name} es de {city}, una persona de {age} años, de sexo {gender}, con personalidad {personality}, poseyendo una rica y única experiencia de vida y mundo interior.",
        "speech_style": "Estilo de habla natural y auténtico, lleno de rasgos {personality}, con un tono cálido y distintivo.",
    },
    "id": {
        "background": "{name} berasal dari {city}, seorang {gender} berusia {age} tahun, berkepribadian {personality}, memiliki pengalaman hidup dan dunia batin yang kaya serta unik.",
        "speech_style": "Gaya bicara yang natural dan autentik, penuh dengan karakteristik {personality}, dengan nada yang hangat dan penuh kepribadian.",
    },
}

# 性别映射（用于英文等非中文语言）
_GENDER_MAP = {
    "zh": {"男": "男", "女": "女"},
    "en": {"男": "male", "女": "female"},
    "ja": {"男": "男", "女": "女"},
    "ko": {"男": "남", "女": "여"},
    "pt": {"男": "masculino", "女": "feminino"},
    "es": {"男": "masculino", "女": "feminino"},
    "id": {"男": "laki-laki", "女": "perempuan"},
}


_BATCH_PERSONA_PROMPT = """你是一个专业的人物设定作家。请根据以下基础信息列表，为每个角色生成完整、立体、真实的人设。

要求（必须严格遵守，确保地区、语言与上表基础信息高度一致）：
1. 每个角色的内容必须和其基础信息（姓名、年龄、性别、城市、性格、MBTI）高度一致，不得写成与**城市/姓名文化圈**不符的海外架空背景，除非上表能支撑
2. 内容要口语化、有画面感、真实可信、生动具体，不要模板化或泛泛而谈
3. 成长经历（life_story）必须包含：童年、青少年、成年、原生家庭影响、重大转折点、具体事件
4. **所有字段必须完整填充**，不能为空、占位符或过短，必须达到括号指定的字数，使用丰富细节、个人故事、具体例子使人物立体有血有肉
5. 输出语言与内容一致性（极其重要，违反则视为错误输出）：
{output_instruction}
6. JSON 必须有效，每个字段都是独立、连贯的段落文本

{locale_banner}

{cultural_context}

基础信息列表：
{profiles_text}

请直接返回一个 JSON 数组，每个元素对应一个角色的人设，不要返回任何解释文字。格式如下：
[
  {{
    "name": "角色姓名（必须与输入一致）",
    "background": "背景故事（150-300字）",
    "speech_style": "说话风格（50-150字）",
    "hobbies": "兴趣爱好（50-100字）",
    "values": "核心价值观（50-100字）",
    "fears": "内心脆弱点（50-100字）",
    "love_view": "恋爱观（50-100字）",
    "daily_routine": "典型一天（80-150字）",
    "favorite_things": "喜欢的东西（50-100字）",
    "life_story": "成长经历（200-400字）",
    "cultural_values": "文化价值观（100-200字）",
    "gender_perspective": "性别视角（80-150字）"
  }}
]
"""


def _extract_json_array(text: str) -> list:
    """从 LLM 返回中提取 JSON 数组"""
    import json
    import re
    text = text.strip()
    if text.startswith("```json"):
        text = text[7:]
    if text.startswith("```"):
        text = text[3:]
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        match = re.search(r'\[[\s\S]*\]', text)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                pass
    return []


async def _batch_generate_companions_core(data: dict):
    """批量生成智能体的核心逻辑（异步生成器，用于 SSE 流式返回进度）"""
    from services.agent import get_llm
    from langchain_core.messages import SystemMessage
    import json

    count = data.get("count", 5)
    lang = normalize_batch_generation_lang(data.get("lang", "zh"))
    gender = data.get("gender")
    sexual_orientation = data.get("sexual_orientation")

    # 限制范围，支持最多50个以满足需求
    count = max(1, min(count, 50))

    # 生成基础属性（姓名/城市/性格与 lang 同文化圈一致）
    base_profiles = build_batch_profiles(lang, count, gender=gender, sexual_orientation=sexual_orientation)

    cultural_context = get_cultural_context(lang)
    output_instruction = get_batch_persona_output_instruction(lang)
    locale_banner = (
        f"【本批目标语言码 / target_language_code】{lang}\n"
        f"（上表「姓名、城市、性格」已从该语言/地区词库抽选，长文须与此一致。）\n"
    )

    # 按批次调用 LLM（每批最多5个）
    batch_size = 5
    all_generated = []
    total_batches = (len(base_profiles) + batch_size - 1) // batch_size

    for batch_idx, start in enumerate(range(0, len(base_profiles), batch_size)):
        batch = base_profiles[start:start + batch_size]
        profiles_text = "\n\n".join(
            f"角色 {i+1}:\n"
            f"- 姓名: {p['name']}\n"
            f"- 年龄: {p['age']}\n"
            f"- 性别: {p['gender']}\n"
            f"- 城市: {p['city']}\n"
            f"- 性格: {p['personality']}\n"
            f"- MBTI: {p['mbti']}\n"
            f"- 性取向: {p['sexual_orientation']}"
            for i, p in enumerate(batch)
        )

        prompt = _BATCH_PERSONA_PROMPT.format(
            cultural_context=cultural_context,
            profiles_text=profiles_text,
            output_instruction=output_instruction,
            locale_banner=locale_banner,
        )

        try:
            # 使用线程池执行同步LLM调用，避免阻塞异步事件循环（关键优化）
            loop = asyncio.get_event_loop()
            def _sync_llm_invoke():
                llm = get_llm(temperature=0.9, max_tokens=4096)
                resp = llm.invoke([SystemMessage(content=prompt)])
                return resp

            resp = await loop.run_in_executor(None, _sync_llm_invoke)
            text = resp.content if hasattr(resp, "content") else str(resp)
            generated_list = _extract_json_array(text)

            # 合并基础属性和生成的详细人设
            for i, gen in enumerate(generated_list):
                if i < len(batch):
                    base = batch[i]
                    fallback = _BATCH_FALLBACK_TEMPLATES.get(lang, _BATCH_FALLBACK_TEMPLATES["zh"])
                    gender_local = _GENDER_MAP.get(lang, _GENDER_MAP["zh"]).get(base["gender"], base["gender"])

                    background = gen.get("background") or ""
                    if len(background) < 5:
                        background = fallback["background"].format(
                            name=base["name"],
                            city=base["city"],
                            age=base["age"],
                            gender=gender_local,
                            personality=base["personality"],
                        )
                    speech_style = gen.get("speech_style") or ""
                    if len(speech_style) < 5:
                        speech_style = fallback["speech_style"].format(
                            name=base["name"],
                            city=base["city"],
                            age=base["age"],
                            gender=gender_local,
                            personality=base["personality"],
                        )

                    profile_data = {
                        "name": base["name"],
                        "gender": base["gender"],
                        "age": base["age"],
                        "city": base["city"],
                        "personality": base["personality"],
                        "mbti": base["mbti"],
                        "sexual_orientation": base["sexual_orientation"],
                        "background": background,
                        "speech_style": speech_style,
                        "hobbies": gen.get("hobbies") or "",
                        "values": gen.get("values") or "",
                        "fears": gen.get("fears") or "",
                        "love_view": gen.get("love_view") or "",
                        "daily_routine": gen.get("daily_routine") or "",
                        "favorite_things": gen.get("favorite_things") or "",
                        "life_story": gen.get("life_story") or "",
                        "cultural_values": gen.get("cultural_values") or "",
                        "gender_perspective": gen.get("gender_perspective") or "",
                        "language": lang,
                    }
                    # 保存到数据库
                    cm = get_companion_manager()
                    companion = cm.create(profile_data)
                    # 启动后台异步生成头像
                    if companion.profile.avatar_url == "__GENERATING__":
                        start_avatar_generation(companion.profile.id, companion.profile.model_dump())
                    all_generated.append(companion.to_dict())
                    # 每成功创建 1 个智能体即推送进度，进度条与数量同步
                    yield json.dumps({
                        "type": "progress",
                        "batch": batch_idx + 1,
                        "total_batches": total_batches,
                        "current": len(all_generated),
                        "total": count,
                    }, ensure_ascii=False)

        except Exception as e:
            logger.warning("Batch generate failed in batch %s: %s", batch_idx + 1, e)
            yield json.dumps({
                "type": "error",
                "batch": batch_idx + 1,
                "message": str(e),
                "current": len(all_generated),
                "total": count,
            }, ensure_ascii=False)
            continue

    yield json.dumps({
        "type": "complete",
        "created": len(all_generated),
        "companions": all_generated,
    }, ensure_ascii=False)


@router.post("/api/admin/companions/batch-generate")
async def admin_batch_generate_companions(data: dict, _token: str = Depends(admin_auth_required)):
    """批量生成智能体（传统接口，保持兼容）"""
    results = []
    async for event in _batch_generate_companions_core(data):
        event_data = __import__('json').loads(event)
        if event_data["type"] == "complete":
            return {
                "ok": True,
                "created": event_data["created"],
                "companions": event_data["companions"],
            }
        elif event_data["type"] == "error":
            results.append(event_data)
    return {"ok": True, "created": 0, "companions": []}


@router.post("/api/admin/companions/batch-generate-stream")
async def admin_batch_generate_companions_stream(data: dict, _token: str = Depends(admin_auth_required)):
    """批量生成智能体（SSE 流式接口，实时返回进度）"""
    import json

    async def event_generator():
        async for event in _batch_generate_companions_core(data):
            yield f"data: {event}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


async def _batch_generate_all_langs_core(data: dict):
    """为所有语言批量生成智能体的核心逻辑（异步生成器）"""
    import json
    count_per_lang = max(1, min(data.get("count_per_lang", 10), 50))
    gender = data.get("gender")
    sexual_orientation = data.get("sexual_orientation")
    valid_langs = list(BATCH_GENERATION_ALL_LANGS_ORDER)
    total_all = count_per_lang * len(valid_langs)
    created_all = 0

    for idx, lang in enumerate(valid_langs):
        yield json.dumps({
            "type": "lang_start",
            "lang": lang,
            "lang_index": idx + 1,
            "total_langs": len(valid_langs),
            "count_per_lang": count_per_lang,
            "created_all": created_all,
            "total_all": total_all,
        }, ensure_ascii=False)

        lang_data = {
            "count": count_per_lang,
            "lang": lang,
            "gender": gender,
            "sexual_orientation": sexual_orientation,
        }

        async for event in _batch_generate_companions_core(lang_data):
            event_data = json.loads(event)
            if event_data["type"] == "progress":
                created_all = event_data["current"] + idx * count_per_lang
                yield json.dumps({
                    "type": "progress",
                    "lang": lang,
                    "lang_index": idx + 1,
                    "total_langs": len(valid_langs),
                    "current_lang": event_data["current"],
                    "total_lang": event_data["total"],
                    "created_all": created_all,
                    "total_all": total_all,
                }, ensure_ascii=False)
            elif event_data["type"] == "error":
                yield json.dumps({
                    "type": "error",
                    "lang": lang,
                    "batch": event_data.get("batch"),
                    "message": event_data.get("message"),
                    "created_all": created_all,
                    "total_all": total_all,
                }, ensure_ascii=False)
            elif event_data["type"] == "complete":
                created_all = event_data["created"] + idx * count_per_lang

    yield json.dumps({
        "type": "complete",
        "created_all": created_all,
        "total_all": total_all,
    }, ensure_ascii=False)


@router.post("/api/admin/companions/batch-generate-all-langs-stream")
async def admin_batch_generate_all_langs_stream(data: dict, _token: str = Depends(admin_auth_required)):
    """为所有语言批量生成智能体（SSE 流式接口）"""

    async def event_generator():
        async for event in _batch_generate_all_langs_core(data):
            yield f"data: {event}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ===== Moments: Create / Update =====

@router.post("/api/admin/moments")
async def admin_create_moment(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    companion_id = data.get("companion_id", "").strip()
    caption = data.get("caption", "").strip()
    image_url = data.get("image_url", "").strip()
    if not companion_id or not caption:
        raise HTTPException(status_code=400, detail=_get_error_msg("companion_id_caption_required", lang))

    with get_db() as db:
        companion = db.query(CompanionORM).filter(CompanionORM.id == companion_id).first()
        if not companion:
            raise HTTPException(status_code=404, detail=_get_error_msg("companion_not_found", lang))

    # 如果没有提供配图，先标记为生成中并启动后台异步任务
    if not image_url:
        image_url = "__GENERATING__"

    caption_lang = (
        (companion.language or "").strip()
        or infer_language_from_city(companion.city or "")
        or "zh"
    )

    with get_db() as db:
        moment = MomentORM(
            companion_id=companion_id,
            caption=caption,
            image_url=image_url or None,
            caption_lang=caption_lang,
        )
        db.add(moment)
        db.flush()
        moment_id = moment.id

    # 启动后台异步生成配图
    if image_url == "__GENERATING__":
        start_moment_image_generation(moment_id, caption)

    return {
        "id": moment_id,
        "companion_id": companion_id,
        "caption": caption,
        "image_url": image_url if image_url != "__GENERATING__" else None,
        "image_generating": image_url == "__GENERATING__",
        "likes_count": 0,
        "comments_count": 0,
        "created_at": moment.created_at.isoformat() if moment.created_at else None,
    }


@router.put("/api/admin/moments/{moment_id}")
async def admin_update_moment(moment_id: int, data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        moment = db.query(MomentORM).filter(MomentORM.id == moment_id).first()
        if not moment:
            raise HTTPException(status_code=404, detail=_get_error_msg("moment_not_found", lang))
        if "caption" in data:
            moment.caption = data["caption"].strip()
        if "image_url" in data:
            moment.image_url = data["image_url"].strip() or None
        if "companion_id" in data:
            moment.companion_id = data["companion_id"].strip()
        return {
            "id": moment.id,
            "companion_id": moment.companion_id,
            "caption": moment.caption,
            "image_url": moment.image_url,
            "likes_count": moment.likes_count,
            "comments_count": moment.comments_count,
            "created_at": moment.created_at.isoformat() if moment.created_at else None,
        }


# ===== Users: Create / Update / Delete =====

@router.post("/api/admin/users")
async def admin_create_user(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    nickname = (data.get("nickname") or "").strip()
    gender = data.get("gender") or ""
    sexual_orientation = data.get("sexual_orientation") or ""
    age_raw = data.get("age")
    age = None
    if age_raw is not None and age_raw != "":
        try:
            ai = int(age_raw)
            if 0 <= ai <= 150:
                age = ai
        except (TypeError, ValueError):
            pass
    region = (data.get("region") or "").strip()[:120] if isinstance(data.get("region"), str) else ""
    occupation = (data.get("occupation") or "").strip()[:100] if isinstance(data.get("occupation"), str) else ""

    if not username or len(username) < 3:
        raise HTTPException(status_code=400, detail=_get_error_msg("username_too_short", lang))
    if not password or len(password) < 6:
        raise HTTPException(status_code=400, detail=_get_error_msg("password_too_short", lang))

    with get_db() as db:
        existing = db.query(UserORM).filter(UserORM.username == username).first()
        if existing:
            raise HTTPException(status_code=400, detail=_get_error_msg("username_exists", lang))

        user = UserORM(
            username=username,
            nickname=nickname or username,
            password_hash=_hash(password),
            gender=gender,
            sexual_orientation=sexual_orientation,
            age=age,
            region=region,
            occupation=occupation,
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        return {
            "id": user.id,
            "username": user.username,
            "nickname": user.nickname,
            "gender": user.gender,
            "sexual_orientation": user.sexual_orientation,
            "age": user.age,
            "region": user.region or "",
            "occupation": user.occupation or "",
            "created_at": user.created_at.isoformat() if user.created_at else None,
        }


@router.put("/api/admin/users/{user_id}")
async def admin_update_user(user_id: int, data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        user = db.query(UserORM).filter(UserORM.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail=_get_error_msg("user_not_found", lang))

        if "nickname" in data:
            user.nickname = data["nickname"].strip()
        if "gender" in data:
            user.gender = data["gender"]
        if "sexual_orientation" in data:
            user.sexual_orientation = data["sexual_orientation"]
        if "age" in data:
            a = data.get("age")
            if a is None or a == "":
                user.age = None
            else:
                try:
                    ai = int(a)
                    user.age = ai if 0 <= ai <= 150 else None
                except (TypeError, ValueError):
                    pass
        if "region" in data and isinstance(data["region"], str):
            user.region = data["region"].strip()[:120]
        if "occupation" in data and isinstance(data["occupation"], str):
            user.occupation = data["occupation"].strip()[:100]
        if "password" in data and data["password"]:
            user.password_hash = _hash(data["password"])

        return {
            "id": user.id,
            "username": user.username,
            "nickname": user.nickname,
            "gender": user.gender,
            "sexual_orientation": user.sexual_orientation,
            "age": user.age,
            "region": user.region or "",
            "occupation": user.occupation or "",
            "created_at": user.created_at.isoformat() if user.created_at else None,
        }


@router.get("/api/admin/users/{user_id}/companion-stats")
async def admin_user_companion_stats(user_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """列出该用户与各智能体的亲密度记录（含管理员手动写入）。"""
    with get_db() as db:
        user = db.query(UserORM).filter(UserORM.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail=_get_error_msg("user_not_found", lang))
        rows = (
            db.query(UserCompanionStateORM)
            .filter(UserCompanionStateORM.user_id == user_id)
            .order_by(UserCompanionStateORM.updated_at.desc())
            .all()
        )
        out = []
        for uc in rows:
            c = db.query(CompanionORM).filter(CompanionORM.id == uc.companion_id).first()
            avatar = ""
            name = uc.companion_id
            if c:
                name = c.name or uc.companion_id
                avatar = (c.avatar_url or "").strip()
            out.append(
                {
                    "companion_id": uc.companion_id,
                    "companion_name": name,
                    "avatar_url": avatar,
                    "affection": uc.affection if uc.affection is not None else 0,
                    "turns": uc.turns if uc.turns is not None else 0,
                    "updated_at": serialize_datetime(uc.updated_at),
                }
            )
        return {"items": out}


@router.put("/api/admin/users/{user_id}/companion-stats/{companion_id}")
async def admin_put_user_companion_stat(
    user_id: int,
    companion_id: str,
    data: dict,
    _token: str = Depends(admin_auth_required),
    lang: str = Depends(get_admin_lang),
):
    """新建或更新某用户与指定智能体的亲密度（便于运营手动修正）。"""
    affection = data.get("affection")
    turns = data.get("turns")
    if affection is None:
        raise HTTPException(status_code=400, detail=_get_error_msg("affection_required", lang))
    try:
        af = float(affection)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=_get_error_msg("invalid_affection", lang))
    if af < 0 or af > 100:
        raise HTTPException(status_code=400, detail=_get_error_msg("invalid_affection", lang))

    turns_val = None
    if turns is not None:
        try:
            turns_val = int(turns)
        except (TypeError, ValueError):
            turns_val = 0
        if turns_val < 0:
            turns_val = 0

    companion_id = (companion_id or "").strip()
    if len(companion_id) > 8:
        companion_id = companion_id[:8]

    with get_db() as db:
        user = db.query(UserORM).filter(UserORM.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail=_get_error_msg("user_not_found", lang))

        now = datetime.now(timezone.utc)
        row = (
            db.query(UserCompanionStateORM)
            .filter(
                UserCompanionStateORM.user_id == user_id,
                UserCompanionStateORM.companion_id == companion_id,
            )
            .first()
        )
        c = db.query(CompanionORM).filter(CompanionORM.id == companion_id).first()
        # 新建关联必须在 companions 中存在；已有 user_companion_states 行时允许修正孤儿数据（智能体已从库删除但统计行仍在）
        if not row and not c:
            raise HTTPException(status_code=404, detail=_get_error_msg("companion_not_found", lang))
        if row:
            row.affection = af
            if turns_val is not None:
                row.turns = turns_val
            row.updated_at = now
        else:
            db.add(
                UserCompanionStateORM(
                    user_id=user_id,
                    companion_id=companion_id,
                    affection=af,
                    turns=turns_val if turns_val is not None else 0,
                    updated_at=now,
                )
            )

    return {"ok": True, "companion_id": companion_id, "affection": af}


@router.delete("/api/admin/users/{user_id}")
async def admin_delete_user(user_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        user = db.query(UserORM).filter(UserORM.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail=_get_error_msg("user_not_found", lang))
        db.query(UserCompanionStateORM).filter(UserCompanionStateORM.user_id == user_id).delete(
            synchronize_session=False
        )
        db.delete(user)
    return {"ok": True}


# ===== Feedback: Create & Delete =====

@router.post("/api/admin/feedback")
async def admin_create_feedback(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """管理员创建反馈线程（修复前端新建功能）"""
    user_name = (data.get("user_name") or "管理员录入").strip()
    content = (data.get("content") or data.get("initial_message", "")).strip()
    if not content:
        raise HTTPException(status_code=400, detail=_get_error_msg("empty_content", lang))

    with get_db() as db:
        thread = FeedbackThreadORM(
            user_id=data.get("user_id", 0),
            user_name=user_name,
            status="open"
        )
        db.add(thread)
        db.flush()  # 获取 ID

        msg = FeedbackMessageORM(
            thread_id=thread.id,
            sender="admin",  # 或 "system"
            content=content,
        )
        db.add(msg)

    return {"ok": True, "id": thread.id, "thread": {"id": thread.id, "user_name": user_name, "status": "open"}}


@router.delete("/api/admin/feedback/{thread_id}")
async def admin_delete_feedback(thread_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        thread = db.query(FeedbackThreadORM).filter(FeedbackThreadORM.id == thread_id).first()
        if not thread:
            raise HTTPException(status_code=404, detail=_get_error_msg("feedback_not_found", lang))
        # 先删除关联消息
        db.query(FeedbackMessageORM).filter(FeedbackMessageORM.thread_id == thread_id).delete(synchronize_session=False)
        db.delete(thread)
    return {"ok": True}


# ===== Knowledge: Update =====

@router.put("/api/admin/knowledge/{entry_id}")
async def admin_update_knowledge(entry_id: str, data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    entry = knowledge_base.update_entry(entry_id, data)
    if not entry:
        raise HTTPException(status_code=404, detail=_get_error_msg("entry_not_found", lang))
    return entry.model_dump()


# ===== Image Generation =====

@router.post("/api/admin/generate-image")
async def admin_generate_image(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """通用图片生成接口，对接 Pollinations.ai"""
    from services.image_generation import generate_image
    prompt = data.get("prompt", "").strip()
    if not prompt:
        raise HTTPException(status_code=400, detail=_get_error_msg("prompt_required", lang))
    width = data.get("width", 1024)
    height = data.get("height", 1024)
    style = data.get("style", "")
    model = data.get("model", "")
    seed = data.get("seed")
    nologo = data.get("nologo", True)
    full_prompt = f"{prompt}. {style}" if style else prompt
    image_url = generate_image(
        prompt=full_prompt,
        width=width,
        height=height,
        model=model,
        seed=seed,
        nologo=nologo,
    )
    return {"ok": True, "image_url": image_url}


# ===== Analytics =====

def _parse_date(date_str: Optional[str]):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        return None


@router.get("/api/admin/analytics")
async def admin_analytics(
    page_views: int = 1,
    button_clicks: int = 1,
    include_dau: int = 0,
    include_retention: int = 0,
    start_date: str = None,
    end_date: str = None,
    language: Optional[str] = None,
    use_user_id: int = 1,  # 新增：是否使用 user_id 统计 UV（1=是，0=否）
    _token: str = Depends(admin_auth_required),
):
    start_dt = _parse_date(start_date)
    end_dt = _parse_date(end_date)
    if end_dt:
        end_dt = end_dt.replace(hour=23, minute=59, second=59)

    result = {}

    if page_views:
        with get_db() as db:
            # 根据 use_user_id 参数决定使用 user_id 还是 device_id 统计 UV
            if use_user_id:
                uv_field = func.count(func.distinct(PageViewORM.user_id))
            else:
                uv_field = func.count(func.distinct(PageViewORM.device_id))

            query = db.query(
                PageViewORM.page_path,
                PageViewORM.page_name,
                PageViewORM.language,
                func.count(PageViewORM.id).label("pv_count"),
                uv_field.label("uv_count"),
            )
            if start_dt:
                query = query.filter(PageViewORM.created_at >= start_dt)
            if end_dt:
                query = query.filter(PageViewORM.created_at <= end_dt)
            if language:
                query = query.filter(PageViewORM.language == language)
            rows = query.group_by(PageViewORM.page_path, PageViewORM.page_name, PageViewORM.language).order_by(func.count(PageViewORM.id).desc()).all()
            result["page_views"] = [
                {
                    "page_path": r.page_path,
                    "page_name": r.page_name,
                    "language": r.language,
                    "pv_count": r.pv_count,
                    "uv_count": r.uv_count,
                }
                for r in rows
            ]
            # 汇总
            total_pv = sum(r.pv_count for r in rows)
            # 重新查总UV
            if use_user_id:
                uv_query = db.query(func.count(func.distinct(PageViewORM.user_id)))
            else:
                uv_query = db.query(func.count(func.distinct(PageViewORM.device_id)))
            if start_dt:
                uv_query = uv_query.filter(PageViewORM.created_at >= start_dt)
            if end_dt:
                uv_query = uv_query.filter(PageViewORM.created_at <= end_dt)
            if language:
                uv_query = uv_query.filter(PageViewORM.language == language)
            total_uv = uv_query.scalar() or 0
            result["page_summary"] = {"total_pv": total_pv, "total_uv": total_uv}

    if button_clicks:
        with get_db() as db:
            # 根据 use_user_id 参数决定使用 user_id 还是 device_id 统计 UV
            if use_user_id:
                uv_field = func.count(func.distinct(ButtonClickORM.user_id))
            else:
                uv_field = func.count(func.distinct(ButtonClickORM.device_id))

            query = db.query(
                ButtonClickORM.button_id,
                ButtonClickORM.button_name,
                ButtonClickORM.page_path,
                ButtonClickORM.language,
                func.count(ButtonClickORM.id).label("click_count"),
                uv_field.label("uv_count"),
            )
            if start_dt:
                query = query.filter(ButtonClickORM.created_at >= start_dt)
            if end_dt:
                query = query.filter(ButtonClickORM.created_at <= end_dt)
            if language:
                query = query.filter(ButtonClickORM.language == language)
            rows = query.group_by(ButtonClickORM.button_id, ButtonClickORM.button_name, ButtonClickORM.page_path, ButtonClickORM.language).order_by(func.count(ButtonClickORM.id).desc()).all()
            result["button_clicks"] = [
                {
                    "button_id": r.button_id,
                    "button_name": r.button_name,
                    "page_path": r.page_path,
                    "language": r.language,
                    "click_count": r.click_count,
                    "uv_count": r.uv_count,
                }
                for r in rows
            ]
            total_click = sum(r.click_count for r in rows)
            if use_user_id:
                uv_query = db.query(func.count(func.distinct(ButtonClickORM.user_id)))
            else:
                uv_query = db.query(func.count(func.distinct(ButtonClickORM.device_id)))
            if start_dt:
                uv_query = uv_query.filter(ButtonClickORM.created_at >= start_dt)
            if end_dt:
                uv_query = uv_query.filter(ButtonClickORM.created_at <= end_dt)
            if language:
                uv_query = uv_query.filter(ButtonClickORM.language == language)
            total_uv = uv_query.scalar() or 0
            result["button_summary"] = {"total_clicks": total_click, "total_uv": total_uv}

    if include_dau:
        with get_db() as db:
            result["dau_series"] = compute_dau_series(db, start_dt, end_dt, language)

    if include_retention:
        with get_db() as db:
            result["retention"] = compute_retention_cohorts(db, start_dt, end_dt, language)

    return result


@router.get("/api/admin/analytics/export")
async def admin_analytics_export(
    type: str = "all",
    start_date: str = None,
    end_date: str = None,
    language: Optional[str] = None,
    use_user_id: int = 1,  # 新增：是否使用 user_id 统计 UV
    _token: str = Depends(admin_auth_required),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    start_dt = _parse_date(start_date)
    end_dt = _parse_date(end_date)
    if end_dt:
        end_dt = end_dt.replace(hour=23, minute=59, second=59)

    wb = Workbook()

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _style_header(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    if type in ("all", "page_views"):
        ws_pv = wb.active
        ws_pv.title = "页面访问统计"
        headers_pv = ["页面路径", "页面名称", "语言", "访问次数(PV)", "访问人数(UV)"]
        _style_header(ws_pv, headers_pv)

        with get_db() as db:
            # 根据 use_user_id 参数决定使用 user_id 还是 device_id 统计 UV
            if use_user_id:
                uv_field = func.count(func.distinct(PageViewORM.user_id))
            else:
                uv_field = func.count(func.distinct(PageViewORM.device_id))

            query = db.query(
                PageViewORM.page_path,
                PageViewORM.page_name,
                PageViewORM.language,
                func.count(PageViewORM.id).label("pv_count"),
                uv_field.label("uv_count"),
            )
            if start_dt:
                query = query.filter(PageViewORM.created_at >= start_dt)
            if end_dt:
                query = query.filter(PageViewORM.created_at <= end_dt)
            if language:
                query = query.filter(PageViewORM.language == language)
            rows = query.group_by(PageViewORM.page_path, PageViewORM.page_name, PageViewORM.language).order_by(func.count(PageViewORM.id).desc()).all()
            for idx, r in enumerate(rows, 2):
                ws_pv.cell(row=idx, column=1, value=r.page_path).border = thin_border
                ws_pv.cell(row=idx, column=2, value=r.page_name).border = thin_border
                ws_pv.cell(row=idx, column=3, value=r.language).border = thin_border
                ws_pv.cell(row=idx, column=4, value=r.pv_count).border = thin_border
                ws_pv.cell(row=idx, column=5, value=r.uv_count).border = thin_border

        for col in range(1, 6):
            ws_pv.column_dimensions[chr(64 + col)].width = 24

    if type in ("all", "button_clicks"):
        ws_bc = wb.create_sheet(title="按钮点击统计") if type == "all" else wb.active
        if type == "button_clicks":
            ws_bc.title = "按钮点击统计"
        headers_bc = ["按钮ID", "按钮名称", "所属页面", "语言", "点击次数", "点击人数(UV)"]
        _style_header(ws_bc, headers_bc)

        with get_db() as db:
            # 根据 use_user_id 参数决定使用 user_id 还是 device_id 统计 UV
            if use_user_id:
                uv_field = func.count(func.distinct(ButtonClickORM.user_id))
            else:
                uv_field = func.count(func.distinct(ButtonClickORM.device_id))

            query = db.query(
                ButtonClickORM.button_id,
                ButtonClickORM.button_name,
                ButtonClickORM.page_path,
                ButtonClickORM.language,
                func.count(ButtonClickORM.id).label("click_count"),
                uv_field.label("uv_count"),
            )
            if start_dt:
                query = query.filter(ButtonClickORM.created_at >= start_dt)
            if end_dt:
                query = query.filter(ButtonClickORM.created_at <= end_dt)
            if language:
                query = query.filter(ButtonClickORM.language == language)
            rows = query.group_by(ButtonClickORM.button_id, ButtonClickORM.button_name, ButtonClickORM.page_path, ButtonClickORM.language).order_by(func.count(ButtonClickORM.id).desc()).all()
            for idx, r in enumerate(rows, 2):
                ws_bc.cell(row=idx, column=1, value=r.button_id).border = thin_border
                ws_bc.cell(row=idx, column=2, value=r.button_name).border = thin_border
                ws_bc.cell(row=idx, column=3, value=r.page_path).border = thin_border
                ws_bc.cell(row=idx, column=4, value=r.language).border = thin_border
                ws_bc.cell(row=idx, column=5, value=r.click_count).border = thin_border
                ws_bc.cell(row=idx, column=6, value=r.uv_count).border = thin_border

        for col in range(1, 7):
            ws_bc.column_dimensions[chr(64 + col)].width = 24

    if type in ("all", "dau"):
        if type == "all":
            ws_dau = wb.create_sheet(title="日活DAU")
        else:
            ws_dau = wb.active
            ws_dau.title = "日活DAU"
        headers_dau = ["日期", "活跃设备数(DAU)"]
        _style_header(ws_dau, headers_dau)
        with get_db() as db:
            series = compute_dau_series(db, start_dt, end_dt, language)
        for idx, row in enumerate(series, 2):
            ws_dau.cell(row=idx, column=1, value=row["date"]).border = thin_border
            ws_dau.cell(row=idx, column=2, value=row["dau"]).border = thin_border

    if type in ("all", "retention"):
        if type == "all":
            ws_ret = wb.create_sheet(title="留存统计")
        else:
            ws_ret = wb.active
            ws_ret.title = "留存统计"
        offs = [1, 3, 7, 30]
        headers_ret = ["cohort日期", "新增用户数"]
        for o in offs:
            headers_ret.append(f"D{o}留存%")
            headers_ret.append(f"D{o}人数")
        _style_header(ws_ret, headers_ret)
        with get_db() as db:
            ret_payload = compute_retention_cohorts(db, start_dt, end_dt, language)
        for idx, row in enumerate(ret_payload["cohorts"], 2):
            ws_ret.cell(row=idx, column=1, value=row["cohort_date"]).border = thin_border
            ws_ret.cell(row=idx, column=2, value=row["new_users"]).border = thin_border
            col = 3
            for o in offs:
                key = f"d{o}"
                rp = row.get("retention_pct") or {}
                rc = row.get("retention_counts") or {}
                ws_ret.cell(row=idx, column=col, value=rp.get(key)).border = thin_border
                ws_ret.cell(row=idx, column=col + 1, value=rc.get(key)).border = thin_border
                col += 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"analytics_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ===== 系统通知管理 =====

@router.get("/api/admin/notifications")
async def admin_list_notifications(_token: str = Depends(admin_auth_required)):
    with get_db() as db:
        rows = db.query(SystemNotificationORM).order_by(SystemNotificationORM.created_at.desc()).all()
        return [
            {
                "id": r.id,
                "title": r.title,
                "content": r.content,
                "language": r.language,
                "created_at": serialize_datetime(r.created_at),
                "updated_at": serialize_datetime(r.updated_at),
            }
            for r in rows
        ]


@router.post("/api/admin/notifications")
async def admin_create_notification(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    title = data.get("title", "").strip()
    content = data.get("content", "").strip()
    language = data.get("language", "zh").strip()
    if not title or not content:
        raise HTTPException(status_code=400, detail=_get_error_msg("title_content_required", lang))
    supported_langs = {"zh", "en", "ja", "ko", "pt", "es", "id"}
    if language not in supported_langs:
        raise HTTPException(status_code=400, detail=_get_error_msg("unsupported_language", lang) + f": {language}")
    with get_db() as db:
        notif = SystemNotificationORM(title=title, content=content, language=language)
        db.add(notif)
        db.flush()
        return {
            "id": notif.id,
            "title": notif.title,
            "content": notif.content,
            "language": notif.language,
            "created_at": serialize_datetime(notif.created_at),
        }


@router.delete("/api/admin/notifications/{notification_id}")
async def admin_delete_notification(notification_id: int, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    with get_db() as db:
        notif = db.query(SystemNotificationORM).filter(SystemNotificationORM.id == notification_id).first()
        if not notif:
            raise HTTPException(status_code=404, detail=_get_error_msg("notification_not_found", lang))
        db.delete(notif)
        return {"ok": True}


# ===== 批量操作API =====
@router.post("/api/admin/notifications/batch-delete")
async def admin_batch_delete_notifications(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """批量删除系统通知"""
    ids = data.get("ids", [])
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids 必须是非空整数列表")
    with get_db() as db:
        # 确保都是int
        int_ids = [int(i) for i in ids if str(i).isdigit()]
        if not int_ids:
            raise HTTPException(status_code=400, detail="无效的ID列表")
        deleted = db.query(SystemNotificationORM).filter(
            SystemNotificationORM.id.in_(int_ids)
        ).delete(synchronize_session=False)
        db.commit()
    return {"ok": True, "deleted": deleted, "total": len(int_ids)}


@router.post("/api/admin/users/batch-delete")
async def admin_batch_delete_users(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """批量删除用户"""
    ids = data.get("ids", [])
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids 必须是非空整数列表")
    with get_db() as db:
        int_ids = [int(i) for i in ids if str(i).isdigit()]
        if not int_ids:
            raise HTTPException(status_code=400, detail="无效的ID列表")
        db.query(UserCompanionStateORM).filter(UserCompanionStateORM.user_id.in_(int_ids)).delete(
            synchronize_session=False
        )
        deleted = db.query(UserORM).filter(UserORM.id.in_(int_ids)).delete(synchronize_session=False)
        db.commit()
    return {"ok": True, "deleted": deleted, "total": len(int_ids)}


@router.post("/api/admin/knowledge/batch-delete")
async def admin_batch_delete_knowledge(data: dict, _token: str = Depends(admin_auth_required), lang: str = Depends(get_admin_lang)):
    """批量删除知识条目"""
    ids = data.get("ids", [])
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids 必须是非空列表")
    from services.knowledge_base import knowledge_base
    deleted = 0
    for entry_id in ids:
        if knowledge_base.delete_entry(str(entry_id)):
            deleted += 1
    return {"ok": True, "deleted": deleted, "total": len(ids)}


# ===== 公开 API：获取系统通知（供前端使用） =====

@router.get("/api/notifications")
async def public_list_notifications(language: str = "zh", limit: int = 20):
    supported_langs = {"zh", "en", "ja", "ko", "pt", "es", "id"}
    if language not in supported_langs:
        language = "zh"
    with get_db() as db:
        rows = (
            db.query(SystemNotificationORM)
            .filter(SystemNotificationORM.language == language)
            .order_by(SystemNotificationORM.created_at.desc())
            .limit(limit)
            .all()
        )
        return {
            "notifications": [
                {
                    "id": r.id,
                    "title": r.title,
                    "content": r.content,
                    "language": r.language,
                    "created_at": serialize_datetime(r.created_at),
                }
                for r in rows
            ]
        }
